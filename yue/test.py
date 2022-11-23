import openpyxl
from xlwt import *
import re
import os
import sys
print("Opening workbook...")
import csv
def get_index_name(id_name):
    wb_id = openpyxl.load_workbook('绩效指标导入模板 (6)(已自动还原).xlsx')
    sheetDataSet = wb_id['DataSheet']
    ids = []
    for i in sheetDataSet['C5':'C447']:
        for j in i:
            if j.value not in ids:
                ids.append(j.value)
    norm_top = ''
    index_name = id_name.split('-')[0].rstrip()
    for id in ids:
        if index_name in id:
            norm_top = id.split("##")[0]
            break
    if norm_top:
        return norm_top
    else:
        print(index_name+" 没有找到")

def get_evaluate_data(file_name):
    # 获取评价表数据
    wb_detail = openpyxl.load_workbook(file_name,data_only = True)
    # sheet_detail = wb_detail[id_name]
    sheet_name = []
    for sheet_name_temp in wb_detail.sheetnames:
        sheet_name.append(sheet_name_temp)
        # break
    # print(sheet_name)
    result = []
    for sub_sheet_name in sheet_name:
        try:
            sheet = wb_detail[sub_sheet_name]
            # if '评价表' not in sheet['A1'].value:
            #     continue
            norm_top = get_index_name(sub_sheet_name)
            
            all_detail = []
            for i in range(3, sheet.max_row):
                sheet_detail = [j.value for j in sheet['B'+str(i):'L'+str(i)][0]]
                all_detail.append(sheet_detail)
            i = 0
            j = 1
            for sub_data in all_detail:
                if not sub_data[0]:
                    continue
                sub_result = {}
                sub_result['norm_id'] = norm_top+'0'+str(j)
                j+=1
                sub_result['norm_name'] = sub_data[1]
                sub_result['norm_type'] = norm_top+"##"+sub_sheet_name.split('-')[0]
                sub_result['usePolicy'] = "全局共享"
                detail_1 = str(handle_text(all_detail[i][9],all_detail[i][4],1))+str(handle_text(all_detail[i][10],all_detail[i][4],1))
                detail_2 = str(handle_text(all_detail[i+1][9],all_detail[i+1][4],2))+'\n' if all_detail[i+1][9] else ''+'\n'
                detail_3 = str(handle_text(all_detail[i+1][10],all_detail[i+1][4],2))+'\n' if all_detail[i+1][10] else ''+'\n'
                detail_4 = str(handle_text(all_detail[i+2][9],all_detail[i+2][4],3)) if all_detail[i+2][9] else ''
                detail_5 = str(handle_text(all_detail[i+2][10],all_detail[i+2][4],3)) if all_detail[i+2][10] else ''
                sub_result['evalStandard'] = str(detail_1+detail_2+detail_3+detail_4+detail_5).rstrip()
                sub_result['indecatorDesc'] = sub_data[2]
                result.append(sub_result)
                i+=3
        except Exception as e:
            # print(traceback.format_exc())
            print(e)
    return result

def get_check_data(file_name):
    # 获取考核卡数据
    wb_detail = openpyxl.load_workbook(file_name,data_only = True)
    sheet_name = []
    for sheet_name_temp in wb_detail.sheetnames:
        sheet_name.append(sheet_name_temp)
    result = []
    for sub_sheet_name in sheet_name:
        try:
            sheet = wb_detail[sub_sheet_name]
            # if '考核卡' not in sheet['A1'].value:
            #     continue
            norm_top = get_index_name(sub_sheet_name)
            all_detail = []
            for i in range(3, sheet.max_row):
                sheet_detail = [j.value for j in sheet['B'+str(i):'E'+str(i)][0]]
                all_detail.append(sheet_detail)
            i = 0
            j = 1
            for sub_data in all_detail:
                if not sub_data[0]:
                    continue
                sub_result = {}
                sub_result['norm_id'] = norm_top+'0'+str(j)
                j+=1
                sub_result['norm_name'] = sub_data[1]
                sub_result['norm_type'] = norm_top+"##"+sub_sheet_name.split('-')[0]
                sub_result['usePolicy'] = "全局共享"
                detail_1 = str(handle_text(all_detail[i][3],1))
                sub_result['evalStandard'] = str(detail_1).rstrip()
                sub_result['indecatorDesc'] = sub_data[2]
                result.append(sub_result)
                i+=1
        except Exception as e:
            print(e)
    return result

def handle_text(content,score_val=0,index=1):
    if not content:
        return ''
    result_content = ''
    index_num = '\d[、,.]\w+'
    score_re = '.*\d分\s*\w*'
    
    for sub_content in content.split('\n'):
        sub_content = sub_content.rstrip().replace(' ','')
        if not sub_content or sub_content == None:
            continue
        if not re.match(index_num,str(sub_content)):
            sub_content = str(index)+'、'+sub_content
        # if not re.match(score_re,str(sub_content)):
        #     sub_content = sub_content+str(score_val)+'分'
        index +=1
        result_content = result_content+sub_content+'\n'
    return result_content

def write_excel(result_list, file_name):
    header = ['norm_id', 'norm_name', 'norm_type', 'usePolicy', 'evalStandard', 'indecatorDesc']
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(title='person',index=0)
    for h_col in range(1, len(header)+1):
        _ = sheet.cell(column=h_col, row=1, value="{0}".format(header[h_col-1]))
    i=2
    for result in result_list:
        j = 1
        for k,v in result.items():
            _ = sheet.cell(row=i,column=j,value="{}".format(v))
            j+=1
        i+=1
    wb.save(filename="{}".format("_evaluate"+file_name.split('/')[-1].split('.')[0])+'.xlsx')
    

def write_csv(result_list, file_name):
    header = ['norm_id', 'norm_name', 'norm_type', 'usePolicy', 'evalStandard', 'indecatorDesc']
    with open('{}.csv'.format(file_name.split('/')[-1].split('.')[0])+'_import', 'w', encoding='utf-8', newline='') as file_obj:
        # 1.创建DicetWriter对象
        dictWriter = csv.DictWriter(file_obj, header)
        # 2.写表头
        dictWriter.writeheader()
        # 3.写入数据(一次性写入多行)
        dictWriter.writerows(result_list)
        
def main():
    # os.chdir('./yue/')
    allfile = os.listdir('file/')
    for file in allfile:
        file_name = "./file/"+file
        result_list = []
        result_list.extend(get_evaluate_data(file_name))
        # result_list.extend(get_check_data(file_name))
        write_excel(result_list,file_name)
        # write_csv(result_list, file_name)
    
if __name__ == "__main__":
    # print(os.getcwd())
    # f = open('./filename.txt', mode='r', encoding='utf-8')
    # f.close()

    main()
    